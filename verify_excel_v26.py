import openpyxl

def verify_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=False)
    sheets = wb.sheetnames
    required_sheets = [
        'Prices', 'SMA64', 'ROC23', 'Eligible', 'Rank',
        'RebalanceDay', 'Status', 'Shares', 'MaxPrice',
        'SL_Trigger', 'Decision', 'EntryRow', 'TradeIndex',
        'Equity', 'Performance', '總績效統計'
    ]

    print(f"Verifying {filepath}...")

    # 1. Check if all sheets exist
    for s in required_sheets:
        if s in sheets:
            print(f"  [OK] Sheet '{s}' exists.")
        else:
            print(f"  [FAIL] Sheet '{s}' is missing!")

    # 2. Check for formulas in key sheets
    ws_stats = wb['總績效統計']
    print(f"Verifying formulas in '總績效統計'...")
    for row in ws_stats.iter_rows(min_row=1, max_row=5, min_col=2, max_col=2):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                 print(f"  [OK] Cell {cell.coordinate} contains formula: {cell.value}")
            else:
                 print(f"  [FAIL] Cell {cell.coordinate} does NOT contain a formula! Value: {cell.value}")

    ws_perf = wb['Performance']
    print(f"Verifying formulas in 'Performance'...")
    # Check first data row of Performance
    for col in range(1, 8):
        cell = ws_perf.cell(row=2, column=col)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
             print(f"  [OK] Cell {cell.coordinate} contains formula.")
        else:
             print(f"  [FAIL] Cell {cell.coordinate} does NOT contain a formula! Value: {cell.value}")

    # 3. Check for specific logic in Decision sheet
    ws_decide = wb['Decision']
    print(f"Verifying formulas in 'Decision' (first stock)...")
    cell_decide = ws_decide.cell(row=67, column=3) # Row 67, Col C (first stock)
    if cell_decide.value and isinstance(cell_decide.value, str) and 'IF(' in cell_decide.value:
         print(f"  [OK] Decision cell contains logic formula.")
    else:
         print(f"  [FAIL] Decision cell does not contain expected logic! Value: {cell_decide.value}")

if __name__ == "__main__":
    verify_excel('trendstrategy_formulas_equity26.xlsx')
