import os
import openpyxl

def verify():
    filename = 'trendstrategy_formulas_equity25-3.xlsx'

    # Check file size (now expecting much larger than 5MB)
    size_mb = os.path.getsize(filename) / (1024 * 1024)
    print(f"File size: {size_mb:.2f} MB")

    # Check sheets
    wb = openpyxl.load_workbook(filename, read_only=True)
    sheets = wb.sheetnames
    print(f"Sheets: {sheets}")
    required_sheets = ['Prices', 'Dashboard', 'Performance', '總績效統計', 'Equity Curve']
    for s in required_sheets:
        if s in sheets:
            print(f"Success: Sheet '{s}' found")
        else:
            print(f"Error: Sheet '{s}' NOT found")

    # Check Performance sheet rows
    ws_perf = wb['Performance']
    print(f"Performance max row: {ws_perf.max_row}")
    if ws_perf.max_row >= 5001:
        print("Success: Performance sheet supports extended trade records")
    else:
        print(f"Error: Performance sheet only has {ws_perf.max_row} rows")

if __name__ == "__main__":
    verify()
