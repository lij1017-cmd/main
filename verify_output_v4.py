import os
import openpyxl

def verify():
    filename = 'trendstrategy_formulas_equity25-4.xlsx'

    # Check file size
    size_mb = os.path.getsize(filename) / (1024 * 1024)
    print(f"File size: {size_mb:.2f} MB")
    if size_mb > 10:
        print("Error: File size exceeds 10MB")
    else:
        print("Success: File size is within 10MB")

    # Check sheets
    wb = openpyxl.load_workbook(filename, read_only=True)
    sheets = wb.sheetnames
    print(f"Sheets: {sheets}")
    required_sheets = ['Prices', 'Dashboard', 'Performance', '總績效統計']
    for s in required_sheets:
        if s in sheets:
            print(f"Success: Sheet '{s}' found")
        else:
            print(f"Error: Sheet '{s}' NOT found")

    # Check Performance sheet rows
    ws_perf = wb['Performance']
    print(f"Performance max row: {ws_perf.max_row}")
    if ws_perf.max_row >= 2001:
        print("Success: Performance sheet supports up to 2000 entries")
    else:
        print(f"Error: Performance sheet only has {ws_perf.max_row} rows")

if __name__ == "__main__":
    verify()
